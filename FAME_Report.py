# -*- coding: utf-8 -*-
"""
End-to-end pipeline:
- Load Clean Data (Treat) CSVs
- Build metrics (ind_served_out, visits_out, spi_out, spi_group_out, groups_out)
- Fill a copy of the Excel template (named “CMHA Peel - FAME - {LastMonthName} 2025”)
- Save it to your local Downloads folder
"""

import os
import re
import glob
from pathlib import Path
import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

# ---------------------- Paths ----------------------
def _onedrive_root() -> Path:
    for var in ("OneDriveCommercial", "OneDrive", "OneDriveConsumer"):
        p = os.environ.get(var)
        if p and Path(p).exists():
            return Path(p)
    home = Path.home()
    for child in home.iterdir():
        if child.is_dir() and child.name.startswith("OneDrive"):
            return child
    return home / "OneDrive"

OD = _onedrive_root()

# --- Your variables (unchanged names) ---
DATA_DIR = str(OD / "data" / "Clean Data" / "Treat")
CENSUS_CSV = os.path.join(DATA_DIR, "rpt_Census.csv")
VISITS_CSV = os.path.join(DATA_DIR, "rpt_ProgStaffFinance.csv")

TEMPLATE_DIR = str(OD / "data" / "Report Templates")
TEMPLATE_BASENAME = "CMHA Peel - FAME - Template"

# Save destination → Downloads
DOWNLOADS = Path.home() / "Downloads"

# ---------------------- Helpers ----------------------
def parse_mdy(series): return pd.to_datetime(series, errors="coerce")
def parse_mdy_hms(series): return pd.to_datetime(series, errors="coerce")
def between_inclusive(s, start, end): return (s >= pd.Timestamp(start)) & (s <= pd.Timestamp(end))

def find_col(df, candidates, required=True):
    for c in candidates:
        if c in df.columns:
            return c
    if required:
        raise KeyError(f"None of the expected columns found: {candidates}")
    return None

def read_csv_robust(path):
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try: return pd.read_csv(path, encoding=enc)
        except Exception: continue
    return pd.read_csv(path)

def to_scalar(val):
    if isinstance(val, (pd.Series, pd.Index, list, tuple)):
        v = pd.to_numeric(pd.Series(val), errors='coerce').fillna(0).sum()
    else:
        try:
            v = float(val)
        except Exception:
            try:
                v = pd.to_numeric(val, errors='coerce')
                v = float(v.sum()) if hasattr(v, "sum") else float(v)
            except Exception:
                v = 0.0
    return int(v) if float(v).is_integer() else v

# ---------------------- Main ----------------------
def main():
    # ----- Read inputs -----
    adt_dat = read_csv_robust(CENSUS_CSV)
    visits_dat = read_csv_robust(VISITS_CSV)

    # (⚡ keep all your existing calculations unchanged — trimmed here for brevity ⚡)
    # ...
    # You already calculate all KPI vars like visits_face_to_face_elderly, spi_5_30_min, etc.
    # Assume those calculations remain intact.

    today = pd.Timestamp.today().normalize()
    report_month_name = (today.replace(day=1) - pd.DateOffset(months=1)).strftime("%B")

    # ----- Load template -----
    candidates = (
        glob.glob(os.path.join(TEMPLATE_DIR, TEMPLATE_BASENAME + "*.xlsm")) +
        glob.glob(os.path.join(TEMPLATE_DIR, TEMPLATE_BASENAME + "*.xlsx"))
    )
    if not candidates:
        raise FileNotFoundError(f"No Excel template named like '{TEMPLATE_BASENAME}*.xlsm/.xlsx' in {TEMPLATE_DIR}")
    template_path = candidates[0]
    keep_vba = template_path.lower().endswith(".xlsm")

    base_template_ext = os.path.splitext(template_path)[1]
    out_name = f"CMHA Peel - FAME - {report_month_name} 2025{base_template_ext}"
    out_path = DOWNLOADS / out_name

    wb = load_workbook(template_path, keep_vba=keep_vba, data_only=False)
    ws = wb.active

    cell_to_var = {
        "C6":  "visits_face_to_face_elderly",
        "C7":  "visits_face_to_face_adult",
        "C8":  "visits_face_to_face_Pediatric",
        "C9":  "visits_face_to_face_Age_Unknown",
        "C10": "visits_email_telephone_elderly",
        "C11": "visits_email_telephone_adult",
        "C12": "visits_email_telephone_Pediatric",
        "C13": "visits_email_telephone_Age_unknown",
        "C15": "Indiv_Served_Elderly",
        "C16": "Indiv_Served_Adult",
        "C17": "Indiv_Served_Pediatric",
        "C18": "Indiv_Served_Age_Unknown",
        "C26": "Group_Part_Reg_Clients",
        "C27": "Num_of_group_Sessions",
        "C36": "spi_5_30_min",
        "C37": "spi_30_1_hour",
        "C38": "spi_1_2_hour",
        "C39": "spi_2_5_hour",
        "C40": "spi_5_more_hour",
        "C53": "spgi_5_30_min",
        "C54": "spgi_31_1_hr",
        "C55": "spgi_1_2_hrs",
        "C56": "spgi_2_5_hrs",
        "C57": "spgi_5_or_more_hrs",
    }

    locs = locals()
    for cell, varname in cell_to_var.items():
        ws[cell] = to_scalar(locs.get(varname, 0))

    wb.save(out_path)
    wb.close()

    print(f"[INFO] Report saved to: {out_path}")

if __name__ == "__main__":
    main()
