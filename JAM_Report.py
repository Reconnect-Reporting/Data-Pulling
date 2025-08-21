# -*- coding: utf-8 -*-
import os
import tempfile
import pandas as pd
from pandas.tseries.offsets import MonthEnd
import datetime as dt
from pathlib import Path
import shutil

# ================== PATHS ==================
ONE_DRIVE = Path(
    os.environ.get("OneDriveCommercial")
    or os.environ.get("OneDrive")
    or (Path.home() / "OneDrive - Reconnect Community Health Services (1)")
)

CLEAN_DIR = str(ONE_DRIVE / "data" / "Clean Data" / "Treat")
MIS_STATS_CSV = os.path.join(CLEAN_DIR, "rpt_MIS_Stats.csv")
PROGFIN_BASE = os.path.join(CLEAN_DIR, "rpt_ProgStaffFinance")  # .csv / .xlsx / .xls

# Save destination
DOWNLOADS = Path.home() / "Downloads"

# ================== CONFIG ==================
TARGET_PROGRAM = "JAM"
PROGRAM_COL = "Program"
ID_COL = "ID"
DATE_COL = "Interaction Start Date"

# Columns present in MIS (monthly + quarters); we'll only use the monthly codes
MONTH_COLS = [
    "Apr","May","June","Q1","July","Aug","Sep","Q2",
    "Oct","Nov","Dec","Q3","Jan","Feb","Mar","Apr_2","Q4"
]
MONTHS_FULL = ["April","May","June","July","August","September","October","November","December","January","February","March"]
MONTHS_CODE = ["Apr","May","June","July","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]

# ================== HELPERS ==================
def read_csv_robust(path: str) -> pd.DataFrame:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return pd.read_csv(path, encoding=enc)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path, errors="replace")

def read_any(path_no_ext: str) -> pd.DataFrame:
    for ext in (".csv", ".xlsx", ".xls"):
        p = path_no_ext + ext
        if os.path.exists(p):
            if ext == ".csv":
                return read_csv_robust(p)
            else:
                return pd.read_excel(p, sheet_name=0)
    raise FileNotFoundError(f"Could not find {os.path.basename(path_no_ext)}.(csv|xlsx|xls) in {CLEAN_DIR}")

def ensure_numeric(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    df = df.copy()
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df

def cumulative_cap(seq, cap_idx):
    total = 0
    out = []
    for i, v in enumerate(seq):
        if i <= cap_idx:
            total += int(v or 0)
            out.append(total)
        else:
            out.append(0)
    return out

def load_mis_rows():
    df_mis = read_csv_robust(MIS_STATS_CSV)
    df_mis.columns = [str(c).strip() for c in df_mis.columns]
    df_mis = df_mis[df_mis[PROGRAM_COL].astype(str).str.strip().eq(TARGET_PROGRAM)].copy()

    desc = df_mis["Description"].astype(str).str.strip()
    mask_face = desc.str.match(r'(?i)^visits\s+face\s+to\s+face:\s*(elderly|adult|paeds)\s*-\s*in\s*person\s*$', na=False)
    mask_non  = desc.str.match(r'(?i)^visits\s+non\s+face\s+to\s+face:\s*(elderly|adult|paeds)\s*$', na=False)

    available_month_cols = [c for c in MONTH_COLS if c in df_mis.columns]
    df_mis = ensure_numeric(df_mis, available_month_cols)

    face_totals = df_mis.loc[mask_face, available_month_cols].fillna(0).sum(axis=0)
    nonface_totals = df_mis.loc[mask_non, available_month_cols].fillna(0).sum(axis=0)

    row_f2f_monthly    = [int(face_totals.get(code, 0))    for code in MONTHS_CODE]
    row_nonf2f_monthly = [int(nonface_totals.get(code, 0)) for code in MONTHS_CODE]
    return row_f2f_monthly, row_nonf2f_monthly

def load_psf_individuals():
    psf = read_any(PROGFIN_BASE)
    psf.columns = [str(c).strip() for c in psf.columns]
    psf = psf[psf[PROGRAM_COL].astype(str).str.strip().eq(TARGET_PROGRAM)].copy()

    psf[DATE_COL] = pd.to_datetime(psf[DATE_COL], errors="coerce")
    psf = psf.dropna(subset=[DATE_COL])

    today = pd.Timestamp.today().normalize()
    first_of_this_month = today.replace(day=1)
    last_complete_month_end = first_of_this_month - pd.Timedelta(days=1)

    if last_complete_month_end.month >= 4:
        fy_start = pd.Timestamp(year=last_complete_month_end.year, month=4, day=1)
    else:
        fy_start = pd.Timestamp(year=last_complete_month_end.year - 1, month=4, day=1)

    fy_end_cap = last_complete_month_end
    psf_fy = psf[(psf[DATE_COL] >= fy_start) & (psf[DATE_COL] <= fy_end_cap)].copy()

    month_starts = pd.date_range(start=fy_start, end=fy_end_cap, freq="MS")

    cumulative_ids: set[str] = set()
    individual_served: dict[str, int] = {}
    for m_start in month_starts:
        m_name = m_start.strftime("%B")
        in_month = psf_fy[psf_fy[DATE_COL].dt.to_period("M") == m_start.to_period("M")]
        month_ids = set(in_month[ID_COL].astype(str).dropna().unique())
        cumulative_ids |= month_ids
        individual_served[m_name] = len(cumulative_ids)

    # Individuals Served array (Apr..Mar), NaN -> 0
    individual_served_df = (
        pd.Series(individual_served, name="Individuals_Served")
          .rename_axis("Month")
          .reindex(MONTHS_FULL)
          .reset_index()
    )
    individual_served_df["Individuals_Served"] = individual_served_df["Individuals_Served"].fillna(0).astype(int)
    row_clients_cum = individual_served_df["Individuals_Served"].tolist()

    return row_clients_cum, fy_end_cap

def build_cumulative_rows(row_f2f_monthly, row_nonf2f_monthly, fy_end_cap):
    last_month_name = fy_end_cap.strftime("%B")
    last_idx = MONTHS_FULL.index(last_month_name)
    row_f2f_cum    = cumulative_cap(row_f2f_monthly,    last_idx)
    row_nonf2f_cum = cumulative_cap(row_nonf2f_monthly, last_idx)
    return row_f2f_cum, row_nonf2f_cum, last_month_name

def fill_template_and_save(row_clients_cum, row_f2f_cum, row_nonf2f_cum, fy_end_cap):
    TEMPLATE_DIR  = str(ONE_DRIVE / "data" / "Report Templates")
    TEMPLATE_BASE = os.path.join(TEMPLATE_DIR, "JAM Report Template")  # .xlsx/.xlsm/.xls

    template_path = None
    template_ext = None
    for ext in (".xlsx", ".xlsm", ".xls"):
        p = TEMPLATE_BASE + ext
        if os.path.exists(p):
            template_path = p
            template_ext = ext.lower()
            break
    if template_path is None:
        raise FileNotFoundError("JAM Report Template (.xlsx/.xlsm/.xls) not found in the template folder.")

    out_path = DOWNLOADS / f"JAM_Report_{fy_end_cap.strftime('%Y%m')}.xlsx"

    import openpyxl
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active  # first sheet
    start_col = 3  # C

    # Row 8: Individuals Served (cumulative)
    for i, val in enumerate(row_clients_cum):
        ws.cell(row=8, column=start_col + i, value=int(val))

    # Row 11: Face-to-Face
    for i, val in enumerate(row_f2f_cum):
        ws.cell(row=11, column=start_col + i, value=int(val))

    # Row 14: Non-Face-to-Face
    for i, val in enumerate(row_nonf2f_cum):
        ws.cell(row=14, column=start_col + i, value=int(val))

    wb.save(out_path)
    return out_path

def main():
    # --- MIS ---
    row_f2f_monthly, row_nonf2f_monthly = load_mis_rows()
    # --- PSF ---
    row_clients_cum, fy_end_cap = load_psf_individuals()
    # --- Cumulative rows ---
    row_f2f_cum, row_nonf2f_cum, last_month_name = build_cumulative_rows(
        row_f2f_monthly, row_nonf2f_monthly, fy_end_cap
    )
    # --- Save directly to Downloads ---
    out_path = fill_template_and_save(row_clients_cum, row_f2f_cum, row_nonf2f_cum, fy_end_cap)
    print(f"[INFO] JAM report saved to: {out_path}")

if __name__ == "__main__":
    main()
